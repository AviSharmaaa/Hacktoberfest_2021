from openpyxl import *
import tkinter as tk
import re


tt_wb = load_workbook(r"timetable.xlsx")
sheet = tt_wb.active

# Add Column Headers
sheet.cell(row=1, column=1).value = "Name"
sheet.cell(row=1, column=2).value = "Email"
sheet.cell(row=1, column=3).value = "Event"
sheet.cell(row=1, column=4).value = "Day"
sheet.cell(row=1, column=5).value = "Time"
sheet.cell(row=1, column=6).value = "Rem-details"

# Tkinter window
root = tk.Tk()
root.title("Add entry to Timetable")
root.geometry("450x200")

name_var = tk.StringVar()
email_var = tk.StringVar()
event_var = tk.StringVar()
day_var = tk.StringVar()
time_var = tk.StringVar()
details_var = tk.StringVar()



# FUNCTIONS cause FUNC it

def check_entry():
	name = name_entry.get()
	email = email_entry.get()
	event = event_entry.get()
	day = day_entry.get()
	time = time_entry.get()
	details = details_entry.get()
	try:
		date_con = 1 <= int(day) <= 7
		exp_time = r"^[0-2][0-3]:[0-5][0-9]$"
		exp_email =  r"^[\w\.\+\-]+\@[\w]+\.[a-z]{2,3}$"
		time_con = bool(re.search(exp_time, time))
		email_con = bool(re.search(exp_email, email))
		return name and email and event and day and time and details and date_con and time_con and email_con
	except:
		return False


def process():
	if check_entry():
		name = name_entry.get()
		email = email_entry.get()
		event = event_entry.get()
		day = day_entry.get()
		time = time_entry.get()
		details = details_entry.get()
		sheet.append((name, email, event, day, time, details))
		tt_wb.save(r"timetable.xlsx")
		
		root.grid_slaves()[0].grid_forget()
		error_label = tk.Label(root, text = "Added!!")
		error_label.grid(row=7, column=1)
	else:
		root.grid_slaves()[0].grid_forget()
		error_label = tk.Label(root, text = "Invalid Input")
		error_label.grid(row=7, column=1)


# Entry boxes
# NAME
name_label = tk.Label(root, text = "Name")
name_entry = tk.Entry(root, textvariable = name_var, width = 40)

# EMAIL
email_label = tk.Label(root, text = "Email")
email_entry = tk.Entry(root, textvariable = email_var, width = 40)

# EVENT
event_label = tk.Label(root, text = "Event")
event_entry = tk.Entry(root, textvariable = event_var, width = 40)

# day
day_label = tk.Label(root, text = "Day(1-7 starting with monday)")
day_entry = tk.Entry(root, textvariable = day_var, width = 40)

# TIME
time_label = tk.Label(root, text = "Time(24H HH:MM)")
time_entry = tk.Entry(root, textvariable = time_var, width = 40)

# DETAILS
details_label = tk.Label(root, text = "Details")
details_entry = tk.Entry(root, textvariable = details_var, width = 40)

# SUBMIT
submit_entry = tk.Button(root, text = "Add Entry", command = process)

# ERROR
error_label = tk.Label(root, text = "Enter details")


name_label.grid(row=0, column=0)
name_entry.grid(row=0, column=1)
email_label.grid(row=1, column=0)
email_entry.grid(row=1, column=1)
event_label.grid(row=2, column=0)
event_entry.grid(row=2, column=1)
day_label.grid(row=3, column=0)
day_entry.grid(row=3, column=1)
time_label.grid(row=4, column=0)
time_entry.grid(row=4, column=1)
details_label.grid(row=5, column=0)
details_entry.grid(row=5, column=1)
submit_entry.grid(row=6, column=1)
error_label.grid(row=7, column=1)


# Save sheet cause dammit
tt_wb.save(r"timetable.xlsx")
root.mainloop()